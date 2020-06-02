# -*- coding: utf-8 -*-
"""
Created on Thu Feb 13 17:19:37 2020

@author: User 1
"""

import openpyxl as xl

from random import randrange

from pyo import *

from time import *

##testing on opening Excel Spreadsheet 
wb = xl.load_workbook('pyoMIDISongDataTest.xlsx')
sheet = wb["Sheet1"]
cell = sheet["a1"]


s = Server().boot()
s.start()
i = eval(input("Enter 1 to play or 0 to stop or 2 to modify: "))
if i == 1:
     
    bFlatMinor = [70, 72, 73, 75, 77, 78, 80, 82] 
    #B-Flat Minor = [A#4/Bb4, C5, C#5/Db5, D#5/Eb5, F5, F#5/Gb5, G#5/Ab5, A#5/Bb5]
    
    midiSongB = []
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        midiSongB.append(cell.value)
        print(cell.value)
    
    #midiSongB =  [70, 72, 75, 73, 78, 73, 77, 75, 78, 75, 80, 78, 80, 78, 77, 75, 72, 70, 73, 77, 78, 75]
    #midiSongB += [72, 75, 82, 80, 82, 80, 78, 77, 72, 73, 72, 68, 66, 60, 63, 60, 61, 58, 60, 65, 66, 68]
    #midiSongB += [70, 72, 73, 80, 77, 82, 84, 87, 89, 92, 94, 92, 89, 84, 80, 77, 73, 75, 72, 73, 72, 70]
    midiSongB += [77, 73, 75, 72, 73, 72, 70]
    
    eMinor = [64, 66, 67, 69, 71, 72, 74, 76]
    #E Minor = [E4, F#4, G4, A4, B4, C5, D5, E5]
    
    midiSongE =  [64, 66, 69, 66, 71, 69, 72, 76, 74, 72, 69, 67, 66, 64, 66, 69, 71, 74, 78, 79]
    midiSongE += [83, 86, 88, 84, 79, 74, 71, 69, 66, 64, 66, 71, 72, 74, 76, 78, 76, 72, 69, 67]
    midiSongE += [69, 72, 76, 78, 83, 84, 83, 81, 79, 78, 76, 72, 74, 71, 69, 71, 67, 69, 66, 64]
    
    
    choice = eval(input("Enter 1 to play with B-Flat Minor or 2 to play with E Minor: "))
    if choice == 1:
        e = Events(freq=EventSeq(midiToHz(midiSongB), occurrences = 0), beat=1.).play()
        midiSong = midiSongB[:]
    if choice == 2:
        e = Events(freq=EventSeq(midiToHz(midiSongE), occurrences = 0), beat=1.).play()
        midiSong = midiSongE[:]
        
    #midiSong to be used with this list, [70, 72, 75, 73, 78, 73, 77, 75, 78, 75, 80, 78, 80, 78, 77, 75, 72, 70, 73, 77, 78, 75, 77, 73, 75, 72, 73, 72, 70]
    #############################################midiSong = midiSong2[:]########
    midiSongHolder = midiSong[:]
    duration = 1
    durationHolder = duration
    restNoteDuration = []
    restNoteDurationHolder = restNoteDuration[:]
    durationHolder2 = duration
while i == 1 or i == 2:
    i = eval(input("Enter 1 to play or 0 to stop or 2 to modify: "))
    if i == 2:
        #midiSong = midiSongHolder 
        
        choice2 = eval(input("Enter 1 to add more note numbers to the song, "
                            "2 to remove a beat from the song, "
                            "3 to modify the duration of all beats from the song, "
                            "4 to increase the octaves of all the notes played from the song, "
                            "5 to decrease the octaves of all the notes played from the song, "
                            "6 to add rest notes in the song in which the total will be the one less than the total beat count, "
                            "7 to modify the song to the original version, "
                            "or 8 to modify the song to the previous version: "))
        e.stop()
        if choice2 == 1:
            midiSongHolder2 = midiSong[:]
            durationHolder2 = duration
            restNoteDurationHolder = restNoteDuration[:]
            numNotesAdded = eval(input("Enter how many notes you want to add: "))
            numNotesAddedHolder = numNotesAdded
            while numNotesAdded > 0:
                noteAdded = eval(input("Enter the note you want to add: "))
                location = eval(input("Enter what location you want the note to be placed: "))
                midiSong.insert(location - 1, noteAdded)
                numNotesAdded -= 1
            if len(restNoteDurationHolder) == 0:
                e = Events(freq=EventSeq(midiToHz(midiSong), occurrences = 0), beat=duration).play()
            else:
                while numNotesAddedHolder > 0:
                    restNoteDuration.append(0)
                    numNotesAddedHolder -= 1
                counter = 0
                while counter < len(midiSong):
                    e = Events(freq=EventSeq(midiToHz([midiSong[counter]]), occurrences = 0), beat=duration).play() 
                    sleep(restNoteDuration[counter])
                    counter += 1 
        if choice2 == 2:
            midiSongHolder2 = midiSong[:]
            durationHolder2 = duration
            restNoteDurationHolder = restNoteDuration[:]
            location = eval(input("Enter what location you want a specfic note to be removed: "))
            del midiSong[location - 1]
            if len(restNoteDurationHolder) == 0:
                e = Events(freq=EventSeq(midiToHz(midiSong), occurrences = 0), beat=duration).play()
            else:
                restNoteDuration.pop()
                counter = 0
                while counter < len(midiSong):
                    e = Events(freq=EventSeq(midiToHz([midiSong[counter]]), occurrences = 0), beat=duration).play() 
                    sleep(restNoteDuration[counter])
                    counter += 1 
        if choice2 == 3:
            #midiDuration = []
            #midiDurationAdded = len(midiSong)
            midiSongHolder2 = midiSong[:]
            durationHolder2 = duration
            restNoteDurationHolder = restNoteDuration[:]
            duration = eval(input("Enter the desired duration you want for all beats of the song: "))
            if len(restNoteDurationHolder) == 0:
                e = Events(freq=EventSeq(midiToHz(midiSong), occurrences = 0), beat=duration).play()
            else:
                counter = 0
                while counter < len(midiSong):
                    e = Events(freq=EventSeq(midiToHz([midiSong[counter]]), occurrences = 0), beat=duration).play() 
                    sleep(restNoteDuration[counter])
                    counter += 1  
        if choice2 == 4:
            midiSongHolder2 = midiSong[:]
            durationHolder2 = duration
            restNoteDurationHolder = restNoteDuration[:]
            while True:
                try:
                    octaveNum = int(input("Enter how many octaves you want each note's octave to be increased: "))
                except ValueError:
                    print("Please enter the correct input again!")
                    continue
                else:
                    break
            counter = 0
            while counter < len(midiSong):
                midiSong[counter] += (octaveNum * 12)
                counter += 1
            if len(restNoteDurationHolder) == 0:
                e = Events(freq=EventSeq(midiToHz(midiSong), occurrences = 0), beat=duration).play()
            else:
                counter = 0
                while counter < len(midiSong):
                    e = Events(freq=EventSeq(midiToHz([midiSong[counter]]), occurrences = 0), beat=duration).play() 
                    sleep(restNoteDuration[counter])
                    counter += 1 
        if choice2 == 5:
            midiSongHolder2 = midiSong[:]
            durationHolder2 = duration
            restNoteDurationHolder = restNoteDuration[:]
            while True:
                try:
                    octaveNum = int(input("Enter how many octaves you want each note's octave to be decreased: "))
                except ValueError:
                    print("Please enter the correct input again!")
                    continue
                else:
                    break
            counter = 0
            while counter < len(midiSong):
                midiSong[counter] -= (octaveNum * 12)
                counter += 1
            if len(restNoteDurationHolder) == 0:
                e = Events(freq=EventSeq(midiToHz(midiSong), occurrences = 0), beat=duration).play()
            else:
                counter = 0
                while counter < len(midiSong):
                    e = Events(freq=EventSeq(midiToHz([midiSong[counter]]), occurrences = 0), beat=duration).play() 
                    sleep(restNoteDuration[counter])
                    counter += 1 
        if choice2 == 6:
            midiSongHolder2 = midiSong[:]
            durationHolder2 = duration
            restNoteDurationHolder = restNoteDuration[:]
            restDurationAdded = len(midiSong) - 1
            counter = 0
            #while restDurationAdded > 0:
            while counter < restDurationAdded:
                durationAdded = eval(input("Enter the desired duration one at a time you want for each rest note for the song (Iteration " + str(counter + 1) + "): "))
                restNoteDuration.append(durationAdded)
                counter += 1
                #restDurationAdded -= 1
            restNoteDuration.append(0)
            counter = 0
            while counter < len(midiSong):
                e = Events(freq=EventSeq(midiToHz([midiSong[counter]]), occurrences = 0), beat=duration).play() 
                sleep(restNoteDuration[counter])
                counter += 1 
        if choice2 == 7:
            midiSong = midiSongHolder[:]
            duration = durationHolder
            e = Events(freq=EventSeq(midiToHz(midiSong), occurrences = 0), beat=duration).play()
        if choice2 == 8:
            midiSong = midiSongHolder2[:]
            duration = durationHolder2
            restNoteDuration = restNoteDurationHolder[:]
            if len(restNoteDurationHolder) == 0:
                e = Events(freq=EventSeq(midiToHz(midiSong), occurrences = 0), beat=duration).play()
            else:
                counter = 0
                while counter < len(midiSong):
                    e = Events(freq=EventSeq(midiToHz([midiSong[counter]]), occurrences = 0), beat=duration).play() 
                    sleep(restNoteDuration[counter])
                    counter += 1 
                
        #midiSong = midiSongHolder[:]
    if i == 0:
        e.stop()
s.stop()